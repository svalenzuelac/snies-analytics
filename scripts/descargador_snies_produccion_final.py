"""
descargador_snies_produccion_final.py
Version produccion final
- Mapeo explícito: articles-ID -> (tipo, año)
- Cruce estudiantes + docentes por IES/año
- Ratio: estudiantes_por_docente = estudiantes / docentes
- Clasificación SUE: 32 universidades verificadas
- CSV limpio: elimina/escapa comas en nombres IES
"""

import subprocess
import sys

print("=" * 70)
print("Instalando dependencias...")
print("=" * 70)

subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", 
                       "pandas", "requests", "openpyxl"])

print("Dependencias instaladas\n")

import requests
import pandas as pd
import openpyxl
import logging
from pathlib import Path
from datetime import datetime
import json
import shutil
import os
import re
import unicodedata
import csv

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DescargadorSNIESProduccion:
    """Cargador ETL para SNIES Analytics"""
    
    MAPEO_ARCHIVOS = {
        "articles-425156_recurso": ("docentes", 2024),
        "articles-425151_recurso": ("estudiantes", 2024),
        "articles-421539_recurso": ("estudiantes", 2023),
        "articles-421822_recurso": ("docentes", 2023),
        "articles-416244_recurso": ("estudiantes", 2022),
        "articles-416249_recurso": ("docentes", 2022),
    }
    
    PATRONES_CANTIDAD = {
        "estudiantes": ["matriculados", "estudiantes", "inscritos", "admitidos"],
        "docentes": ["no. de docentes", "numero de docentes", "docentes", "no docentes"],
    }
    
    UNIVERSIDADES_SUE = {
        "universidad atlántico",
        "universidad de cartagena",
        "universidad popular del cesar",
        "universidad de córdoba",
        "universidad de la guajira",
        "universidad de magdalena",
        "universidad de sucre",
        "universidad distrital francisco josé de caldas",
        "universidad distrital-francisco jose de caldas",
        "universidad colegio mayor de cundinamarca",
        "universidad-colegio mayor de cundinamarca",
        "universidad militar nueva granada",
        "universidad militar-nueva granada",
        "universidad pedagógica nacional",
        "universidad pedagogica nacional",
        "universidad nacional de colombia",
        "universidad pedagógica y tecnológica de colombia",
        "universidad de cundinamarca",
        "universidad nacional abierta y a distancia",
        "unad",
        "universidad del tolima",
        "universidad surcolombiana",
        "universidad de antioquia",
        "universidad tecnológica de pereira",
        "universidad del quindío",
        "universidad de caldas",
        "universidad del valle",
        "universidad del cauca",
        "universidad de nariño",
        "universidad del pacífico",
        "universidad de la amazonia",
        "universidad autónoma intercultural indígena",
        "universidad tecnológica del chocó",
        "universidad de los llanos",
        "universidad de pamploma",
        "universidad industrial de santander",
        "universidad francisco de paula santander",
    }
    
    URLS_DESCARGA = {
        nombre: f"https://snies.mineducacion.gov.co/1778/{nombre}.xlsx"
        for nombre in MAPEO_ARCHIVOS.keys()
    }
    
    def __init__(self, directorio_salida="data"):
        self.directorio_salida = Path(directorio_salida)
        self._limpiar_directorio_seguro()
        self.directorio_salida.mkdir(parents=True, exist_ok=True)
        
        self.sesion = requests.Session()
        self.sesion.headers.update({'User-Agent': 'SNIES-Analytics/1.0'})
        
        self.datos_estudiantes = {}
        self.datos_docentes = {}
        self.datos_sector = {}
        
        self.metadatos = {
            "marca_tiempo": datetime.now().isoformat(),
            "descargas": {},
            "procesados": {}
        }
        
        self.archivo_salida = self.directorio_salida / "snies_relacion_estudiante_docente.csv"
    
    def _limpiar_directorio_seguro(self):
        if not self.directorio_salida.exists():
            return
        
        logger.info("Limpiando directorio anterior...")
        try:
            shutil.rmtree(self.directorio_salida)
            logger.info("Borrado exitoso")
        except PermissionError:
            logger.warning("Permisos: intentando borrado selectivo")
            try:
                for raiz, carpetas, archivos in os.walk(self.directorio_salida, topdown=False):
                    for archivo in archivos:
                        try:
                            os.remove(os.path.join(raiz, archivo))
                        except:
                            pass
                    for nombre_carpeta in carpetas:
                        try:
                            os.rmdir(os.path.join(raiz, nombre_carpeta))
                        except:
                            pass
                try:
                    os.rmdir(self.directorio_salida)
                except:
                    pass
                logger.info("Borrado selectivo completado")
            except:
                logger.warning("Continuando...")
    
    def detectar_hoja_datos(self, archivo_excel):
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        excluir = ['índice', 'indice', 'sumario', 'contenido', 'tabla', 'resumen']
        
        for nombre_hoja in wb.sheetnames:
            nombre_bajo = nombre_hoja.lower()
            if not any(x in nombre_bajo for x in excluir):
                return nombre_hoja
        
        return wb.sheetnames[0] if wb.sheetnames else None
    
    def normalizar_encabezado(self, nombre_col):
        normalizado = str(nombre_col).replace('\n', ' ').strip()
        normalizado = re.sub(r'\s+', ' ', normalizado)
        return normalizado
    
    def encontrar_columna_por_patron(self, df, patrones):
        if isinstance(patrones, str):
            patrones = [patrones]
        
        patrones = [p.lower() for p in patrones]
        
        for patron in patrones:
            for col in df.columns:
                col_norm = self.normalizar_encabezado(col).lower()
                palabras = patron.split()
                if all(palabra in col_norm for palabra in palabras):
                    return col
        
        return None
    
    def normalizar_departamento(self, valor):
        if pd.isna(valor):
            return ""
        
        s = str(valor).lower().strip()
        s = ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')
        s = s.replace('d.c.', '').replace('dc', '').strip()
        s = s.replace(',', '').replace('.', '').strip()
        s = re.sub(r'\s+', '', s)
        
        return s
    
    def limpiar_nombre_ies(self, nombre):
        if pd.isna(nombre):
            return ""
        
        nombre = str(nombre).strip()
        nombre = nombre.replace(',', '')
        nombre = re.sub(r'\s+', ' ', nombre).strip()
        nombre = nombre.rstrip('-').strip()
        return nombre
    
    def normalizar_para_sue(self, texto):
        if pd.isna(texto):
            return ""
        
        texto = str(texto).lower().strip()
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                       if unicodedata.category(c) != 'Mn')
        texto = texto.replace('-', ' ')
        texto = re.sub(r'\s+', ' ', texto).strip()
        
        return texto
    
    def es_sue(self, nombre_ies):
        nombre_norm = self.normalizar_para_sue(nombre_ies)
        
        for sue in self.UNIVERSIDADES_SUE:
            sue_norm = self.normalizar_para_sue(sue)
            
            if nombre_norm == sue_norm:
                return True
            
            if sue_norm in nombre_norm or nombre_norm in sue_norm:
                return True
        
        return False
    
    def leer_excel_produccion(self, archivo_excel, nombre, tipo_esperado, ano_esperado):
        logger.info("Leyendo: " + nombre)
        logger.info("Tipo/Año: " + str(tipo_esperado) + "/" + str(ano_esperado))
        
        try:
            hoja = self.detectar_hoja_datos(archivo_excel)
            logger.info("Hoja: " + str(hoja))
            
            df_temp = pd.read_excel(archivo_excel, sheet_name=hoja, nrows=10, header=None)
            
            fila_encabezado = 7
            for idx, fila in df_temp.iterrows():
                fila_str = ' '.join(str(x).lower() for x in fila if pd.notna(x))
                if ('institución' in fila_str or 'código' in fila_str) and 'departamento' in fila_str:
                    fila_encabezado = idx
                    logger.info("Encabezado en fila " + str(idx + 1))
                    break
            
            df = pd.read_excel(archivo_excel, sheet_name=hoja, skiprows=fila_encabezado)
            df = df.dropna(how='all')
            df = df.reset_index(drop=True)
            
            tamaño_mb = Path(archivo_excel).stat().st_size / (1024 * 1024)
            logger.info(str(len(df)) + " registros (" + str(round(tamaño_mb, 2)) + " MB)")
            
            self.metadatos["descargas"][nombre] = {
                "registros": len(df),
                "tipo": tipo_esperado,
                "ano": ano_esperado,
                "hoja": hoja
            }
            
            return df, tipo_esperado, ano_esperado
        
        except Exception as e:
            logger.error("Error: " + str(e))
            import traceback
            traceback.print_exc()
            return None, None, None
    
    def procesar_y_almacenar(self, df, tipo, ano, nombre):
        if df is None or len(df) == 0:
            return
        
        logger.info("Procesando " + tipo + " " + str(ano))
        
        df = df.copy()
        df.columns = [self.normalizar_encabezado(str(c)).lower() for c in df.columns]
        
        col_depto = self.encontrar_columna_por_patron(df, "departamento domicilio ies")
        if col_depto is None:
            logger.warning("Columna de departamento no encontrada")
            return
        
        df['_depto_norm'] = df[col_depto].apply(self.normalizar_departamento)
        mascara = df['_depto_norm'].str.contains('bogota', na=False)
        df_filtrado = df[mascara]
        df_filtrado = df_filtrado.drop('_depto_norm', axis=1)
        
        logger.info("Registros Bogotá: " + str(len(df_filtrado)))
        
        if len(df_filtrado) == 0:
            logger.warning("Sin registros de Bogotá")
            return
        
        col_ies = self.encontrar_columna_por_patron(df_filtrado, "institución educación superior ies")
        if col_ies is None:
            col_ies = self.encontrar_columna_por_patron(df_filtrado, "institución educación superior")
        
        col_sector = self.encontrar_columna_por_patron(df_filtrado, "sector ies")
        col_cantidad = self.encontrar_columna_por_patron(df_filtrado, self.PATRONES_CANTIDAD.get(tipo, []))
        
        if col_ies is None or col_cantidad is None:
            logger.error("Columnas no encontradas")
            return
        
        logger.info("Columnas encontradas")
        
        df_filtrado[col_cantidad] = pd.to_numeric(df_filtrado[col_cantidad], errors='coerce').fillna(0)
        
        agg = df_filtrado.groupby(col_ies)[col_cantidad].sum().reset_index()
        agg.columns = [col_ies, 'total_' + tipo]
        
        for idx, fila in agg.iterrows():
            nombre_ies_crudo = fila[col_ies]
            nombre_ies = self.limpiar_nombre_ies(nombre_ies_crudo)
            total = int(fila['total_' + tipo])
            
            clave = (nombre_ies, ano)
            if tipo == "estudiantes":
                self.datos_estudiantes[clave] = total
            else:
                self.datos_docentes[clave] = total
            
            if col_sector:
                valor_sector = df_filtrado[df_filtrado[col_ies] == nombre_ies_crudo][col_sector].iloc[0]
                self.datos_sector[nombre_ies] = valor_sector
        
        logger.info("IES procesadas: " + str(len(agg)))
        logger.info("Total " + tipo + ": " + str(int(agg['total_' + tipo].sum())))
    
    def generar_resultado_final(self):
        logger.info("=" * 70)
        logger.info("GENERANDO RESULTADO FINAL")
        logger.info("=" * 70)
        
        todas_las_ies = set()
        todos_los_anos = set()
        
        for (ies, ano) in self.datos_estudiantes.keys():
            todas_las_ies.add(ies)
            todos_los_anos.add(ano)
        
        for (ies, ano) in self.datos_docentes.keys():
            todas_las_ies.add(ies)
            todos_los_anos.add(ano)
        
        logger.info("IES únicas: " + str(len(todas_las_ies)))
        logger.info("Años: " + str(sorted(todos_los_anos)))
        
        resultados = []
        
        for nombre_ies in sorted(todas_las_ies):
            for ano in sorted(todos_los_anos):
                clave = (nombre_ies, ano)
                
                estudiantes = self.datos_estudiantes.get(clave, 0)
                docentes = self.datos_docentes.get(clave, 0)
                
                if docentes > 0:
                    ratio = round(estudiantes / docentes, 2)
                else:
                    ratio = None
                
                es_sue = self.es_sue(nombre_ies)
                sector_sue = "SUE" if es_sue else "No SUE"
                
                sector = self.datos_sector.get(nombre_ies, "")
                
                resultados.append({
                    'nombre_ies': nombre_ies,
                    'año': ano,
                    'total_estudiantes': estudiantes,
                    'total_docentes': docentes,
                    'estudiantes_por_docente': ratio,
                    'sector_ies': sector,
                    'clasificacion_sue': sector_sue
                })
        
        df_resultado = pd.DataFrame(resultados)
        df_resultado = df_resultado.sort_values(['año', 'nombre_ies'])
        
        df_resultado.to_csv(
            self.archivo_salida, 
            index=False, 
            encoding='utf-8',
            quoting=csv.QUOTE_NONNUMERIC,
            doublequote=True
        )
        
        logger.info("RESULTADO FINAL:")
        logger.info("Registros: " + str(len(df_resultado)))
        logger.info("IES únicas: " + str(df_resultado['nombre_ies'].nunique()))
        logger.info("Años: " + str(sorted(df_resultado['año'].unique())))
        logger.info("Archivo: " + str(self.archivo_salida))
        
        logger.info("Primeros 5 registros:")
        logger.info(df_resultado.head(5).to_string())
        
        return df_resultado
    
    def descargar_archivo(self, url, nombre, max_reintentos=2):
        logger.info("Descargando: " + nombre)
        
        for intento in range(max_reintentos):
            try:
                respuesta = self.sesion.get(url, timeout=30)
                respuesta.raise_for_status()
                
                archivo_salida = self.directorio_salida / (nombre + ".xlsx")
                with open(archivo_salida, 'wb') as f:
                    f.write(respuesta.content)
                
                tamaño_mb = len(respuesta.content) / (1024 * 1024)
                logger.info("Descargado (" + str(round(tamaño_mb, 2)) + " MB)")
                
                return archivo_salida
            
            except Exception as e:
                logger.warning("Intento " + str(intento + 1) + " falló: " + str(e)[:60])
                if intento == max_reintentos - 1:
                    logger.error("No disponible")
                    return None
        
        return None
    
    def descargar_y_procesar(self):
        logger.info("=" * 70)
        logger.info("DESCARGADOR SNIES PRODUCCION FINAL")
        logger.info("=" * 70)
        logger.info("Relación Estudiante/Docente + Clasificación SUE")
        logger.info("Período: 2022-2024 | Alcance: Bogotá")
        logger.info("CSV: LIMPIO (sin comas en nombres)")
        
        logger.info("=" * 70)
        logger.info("FASE 1: DESCARGAS")
        logger.info("=" * 70)
        
        archivos_descargados = {}
        
        for nombre, url in self.URLS_DESCARGA.items():
            archivo = self.descargar_archivo(url, nombre)
            if archivo:
                archivos_descargados[nombre] = archivo
        
        logger.info("=" * 70)
        logger.info("FASE 2: PROCESAMIENTO")
        logger.info("=" * 70)
        
        for nombre, archivo in archivos_descargados.items():
            tipo, ano = self.MAPEO_ARCHIVOS[nombre]
            df, tipo, ano = self.leer_excel_produccion(archivo, nombre, tipo, ano)
            if df is not None:
                self.procesar_y_almacenar(df, tipo, ano, nombre)
        
        logger.info("=" * 70)
        logger.info("FASE 3: GENERACIÓN DE RESULTADO")
        logger.info("=" * 70)
        
        df_final = self.generar_resultado_final()
        
        self.metadatos["estado"] = "completado"
        archivo_metadatos = self.directorio_salida / "metadatos_descarga.json"
        with open(archivo_metadatos, 'w', encoding='utf-8') as f:
            json.dump(self.metadatos, f, indent=2, ensure_ascii=False)
        
        logger.info("Metadatos: " + str(archivo_metadatos))
        
        logger.info("=" * 70)
        logger.info("PROCESAMIENTO COMPLETADO")
        logger.info("=" * 70)
        logger.info("Próximo:")
        logger.info("  docker-compose down -v")
        logger.info("  docker-compose up")

def main():
    descargador = DescargadorSNIESProduccion(directorio_salida="data")
    descargador.descargar_y_procesar()

if __name__ == "__main__":
    main()
